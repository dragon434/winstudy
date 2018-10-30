#!/usr/bin/env python
# -*- coding: UTF-8 -*-
##########################
# dsc : 统计puv量
#       1. 转换exce
#       2. 转换html
# version : v1.0
# date : 2018-09-18
# dsc : 添加时间段统计
# version : v1.2
# date : 2018-09-27
# author : @jiawenlong
##########################

from __future__ import unicode_literals
from pyecharts import Bar, Page, Line
from datetime import datetime
from datetime import timedelta
from openpyxl import Workbook
from openpyxl import load_workbook
import xlrd
import json
import os
import sys


def give_time():
    today = datetime.today()
    today_date = datetime.date(today) - timedelta(days=1)
    return today_date.strftime("%Y%m%d"), today_date.strftime("%Y%m"), today_date.strftime("%d")


def get_puv(file):
    puv = open(file, 'r')
    data = puv.readlines()
    PUV = {}
    for lines in data:
        line = lines.strip('\n')
        js = json.loads(line)
        PUV.update(js)
    return PUV


def add_head(sheet):
    head = ['', '日期', 'PV', 'UV', "IP"]
    for i in range(1, 5):
        s = sheet.cell(row=1, column=i)
        s.value = head[i]
    return sheet


def get_days(sheet):
    for i in range(2, 31):
        s = sheet.cell(row=i, column=1)
        if s.value:
            continue
        else:
            return sheet.cell(row=i - 1, column=1).value, i


def add_data(sheet, day, key, value, hang):
    rc_1 = sheet.cell(row=hang, column=1)
    pv = sheet.cell(row=1, column=2).value
    uv = sheet.cell(row=1, column=3).value
    ip = sheet.cell(row=1, column=4).value

    rc_1.value = day
    if pv == key:
        sheet.cell(row=hang, column=2).value = value
    if uv == key:
        sheet.cell(row=hang, column=3).value = value
    if ip == key:
        sheet.cell(row=hang, column=4).value = value

    return sheet


def add_data2sheet(sheet, puv_dic, day, sheet_name):
    worktable = sheet
    add_head(worktable)
    hnum_value, hang = get_days(worktable)

    for key in puv_dic:
        for keys, values in puv_dic[key].items():
            if sheet_name == key:
                add_data(worktable, day, keys, values, hang)
            else:
                break


def export_excel(file_name, names, puv_dic, day):

    if os.path.isfile(file_name):
        puv_exc = load_workbook(file_name)
        # 删除 多余的 sheet 表， puv_exc.sheetnames 以列表形式返回工作簿所有sheet表名
        if 'Sheet' in puv_exc.sheetnames:
            del puv_exc['Sheet']
        for name in names:
            worktable = puv_exc[name]
            # 获取 sheet 表名
            title = worktable.title
            add_data2sheet(worktable, puv_dic, day, title)
    else:
        puv_exc = Workbook()
        for name in names:
            worktable = puv_exc.create_sheet(name)
            add_data2sheet(worktable, puv_dic, day, name)

    puv_exc.save(file_name)


def add_html(m, d):
    data = xlrd.open_workbook('/mnt/data/bank/section/account_ekeguan_com' + m + '.xlsx')
    www = xlrd.open_workbook('/mnt/data/bank/section/www_ekeguan_com' + m + '.xlsx')
    table = data.sheets()[0]
    bank = www.sheets()[0]

    CLOTHES = table.col_values(0)
    clothes_pv = table.col_values(1)
    clothes_uv = table.col_values(2)
    clothes_ip = table.col_values(3)

    bar = Line("account_ekeguan_com_" + d + "号", height=350)
    bar.add("PV", CLOTHES, clothes_pv)
    bar.add("UV", CLOTHES, clothes_uv)
    bar.add("IP", CLOTHES, clothes_ip, is_stack=False, is_datazoom_show=True)

    CLOTHES = bank.col_values(0)
    clothes_pv = bank.col_values(1)
    clothes_uv = bank.col_values(2)
    clothes_ip = bank.col_values(3)

    line = Line("www_ekeguan_com_" + d + "号", title_top='1%')
    line.add("PV", CLOTHES, clothes_pv)
    line.add("UV", CLOTHES, clothes_uv)
    line.add("IP", CLOTHES, clothes_ip, is_stack=False, is_datazoom_show=True)

    return bar, line


def create_html(excelfile, html, day, m):
    page = Page()

    data = xlrd.open_workbook(excelfile)
    table = data.sheets()[0]
    bank = data.sheets()[1]

    CLOTHES = table.col_values(0)
    clothes_pv = table.col_values(1)
    clothes_uv = table.col_values(2)
    clothes_ip = table.col_values(3)

    bar = Bar("account_ekeguan_com", height=350)
    bar.add("PV", CLOTHES, clothes_pv)
    bar.add("UV", CLOTHES, clothes_uv)
    bar.add("IP", CLOTHES, clothes_ip, is_stack=False, is_datazoom_show=True)

    CLOTHES = bank.col_values(0)
    clothes_pv = bank.col_values(1)
    clothes_uv = bank.col_values(2)
    clothes_ip = bank.col_values(3)

    line = Bar("www_ekeguan_com", title_top='1%')
    line.add("PV", CLOTHES, clothes_pv)
    line.add("UV", CLOTHES, clothes_uv)
    line.add("IP", CLOTHES, clothes_ip, is_stack=False, is_datazoom_show=True)

    day_bar, day_line = add_html(m, day)
    page.add(bar)
    page.add(line)
    page.add(day_bar)
    page.add(day_line)
    page.render(html)


if __name__ == '__main__':
    # sys.argv 是列表
    # sys.argv[2:] 获取 除脚本名称 和项目名称外的所有参数
    param = sys.argv[2:]
    print("param:", param)
    pro = sys.argv[1]
    date_ymd, date_ym, date_day = give_time()
    soruce_file = "E:\lrzsz\sz\\" + pro + "\\" + date_ymd + '.json'
    excel_file = "E:\lrzsz\sz\\" + pro + "\\" + date_ym + ".xlsx"
    html_file = "E:\lrzsz\sz\\" + pro + "\\" + pro + date_ym + ".html"
    # print soruce_file,excel_file,html_file
    #
    data_dic = get_puv(soruce_file)
    export_excel(excel_file, param, data_dic, date_day)
    # create_html(excel_file, html_file, date_day, date_ym)
