#!/usr/bin/python
# -*- coding: UTF-8 -*-


from datetime import datetime
from datetime import timedelta
from openpyxl import Workbook
from openpyxl import load_workbook
import json
import os


def give_time():
    """ last_date 获取的就是前一天的日期
        例如： 今天是2018-10-01，那么 today_date 就是 2018-09-30
        today.day 是这个月的第几天，不是几号 可以直接进行加减
        arrow 可以直接获取日期 前一个月的日期
        当前状态下，返回的 日期 都是字符串
    """
    today = datetime.today()
    last_date = datetime.date(today) - timedelta(days=1)
    return last_date.strftime("%Y%m%d"), last_date.strftime("%Y%m"), last_date.strftime('%d')


# 返回数值字典
def get_puv(file):
    puv = open(file, 'r')
    data = puv.readlines()
    PUV = {}
    for lines in data:
        line = lines.strip('\n')
        js = json.loads(line)
        PUV.update(js)
    return PUV


# 添加excel第一行固定值
def add_head(sheet):
    head = ['', '时间小时', 'pv', 'uv', "ip"]
    for i in range(1, 5):
        s = sheet.cell(row=1, column=i)
        s.value = head[i]
    # return sheet


# 获取有数据的行 的下一行行号
def get_days(sheet):
    for i in range(2, 26):
        s = sheet.cell(row=i, column=1)
        # print(s.value)
        if s.value:
            continue
        else:
            return i
            # return sheet.cell(row=i-1, column=1).value, i


# 把数据写入内存中的excel对应的行列
def add_data(sheet, kk, vkey, value, hang):
    pv = sheet.cell(row=1, column=2).value
    uv = sheet.cell(row=1, column=3).value
    ip = sheet.cell(row=1, column=4).value

    sheet.cell(row=hang, column=1).value = vkey

    if pv == kk:
        sheet.cell(row=hang, column=2).value = value
    if uv == kk:
        sheet.cell(row=hang, column=3).value = value
    if ip == kk:
        sheet.cell(row=hang, column=4).value = value
    # return sheet


# 把原始数据写入文件
def export_excel(file_name, puv_dic, day):
    if os.path.isfile(file_name):
        puv_exc = load_workbook(file_name)
        worktable1 = puv_exc.create_sheet(day, 0)
    else:
        puv_exc = Workbook()
        worktable1 = puv_exc.create_sheet(day, 0)

    add_head(worktable1)
    hang_num = get_days(worktable1)

    for key in puv_dic:
        for keys, values in puv_dic[key].items():
            print(keys, values, hang_num)
            add_data(worktable1, key, keys, values, hang_num)
            if hang_num < 25:
                hang_num += 1
            else:
                hang_num = 2

    puv_exc.save(file_name)


date_ymd, date_ym, date_day = give_time()
soruce_file = "E:\lrzsz\sz\\" + date_ymd + '.json'
data_dic = get_puv(soruce_file)
for key in data_dic:
    excel_file = "E:\lrzsz\sz\\" + key + date_ym + ".xlsx"
    export_excel(excel_file, data_dic[key], date_day)
