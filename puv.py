#!/usr/bin/python
# -*- coding: UTF-8 -*-
#
from datetime import datetime
from datetime import timedelta
from openpyxl import Workbook
from openpyxl import load_workbook
import json
import os


def give_time():
    today = datetime.today()
    today_date = datetime.date(today) - timedelta(days=1)
    return today_date.strftime("%Y%m%d"), today_date.strftime("%Y%m"), today.day - 1

# DATE, DAY = give_time()
# print(DATE, DAY)


# 返回数值字典
def get_puv(file):
    puv = open(file, 'r')
    data = puv.readlines()
    PUV = {}
    for lines in data:
        line = lines.strip('\n')
        js = json.loads(line)
        PUV.update(js)
    return PUV


# 添加第一行固定值
def add_head(sheet):
    head = ['', '日期', 'PV', 'UV', "IP"]
    for i in range(1, 5):
        s = sheet.cell(row=1, column=i)
        s.value = head[i]
    return sheet


# 获取有数据的行的下一行行号
def get_days(sheet):
    for i in range(2, 31):
        s = sheet.cell(row=i, column=1)
        if s.value:
            continue
        else:
            return sheet.cell(row=i-1, column=1).value, i


# 把数据写入内存中的excel对应的行列
def add_data(sheet, day, key, value, hang):
    rc_1 = sheet.cell(row=hang, column=1)
    pv = sheet.cell(row=1, column=2).value
    uv = sheet.cell(row=1, column=3).value
    ip = sheet.cell(row=1, column=4).value

    rc_1.value = day
    if pv == key:
        sheet.cell(row=hang, column=2).value = value
    if uv == key:
        sheet.cell(row=hang, column=3).value = value
    if ip == key:
        sheet.cell(row=hang, column=4).value = value

    return sheet


# 把原始数据写入文件
def export_excel(file_name, puv_dic, day):
    if os.path.isfile(file_name):
        puv_exc = load_workbook(file_name)
        worktable1 = puv_exc.active
        worktable2 = puv_exc['www_ekeguan_com']
    else:
        puv_exc = Workbook()
        worktable1 = puv_exc.create_sheet('account_ekeguan_com', 0)
        worktable2 = puv_exc.create_sheet('www_ekeguan_com', 1)

    add_head(worktable1)
    add_head(worktable2)
    hnum_value1, hang1 = get_days(worktable1)
    hnum_value2, hang2 = get_days(worktable2)

    for key in puv_dic:
        for keys, values in puv_dic[key].items():
            if key == 'account_ekeguan_com':
                add_data(worktable1, day, keys, values, hang1)
            if key == 'www_ekeguan_com':
                add_data(worktable2, day, keys, values, hang2)

    puv_exc.save(file_name)


date_ymd, date_ym, date_day = give_time()
soruce_file = "E:\lrzsz\sz\\" + date_ymd
excel_file = "E:\lrzsz\sz\\" + date_ym + ".xlsx"
data_dic = get_puv(soruce_file)
export_excel(excel_file, data_dic, date_day)




